
import openpyxl as xl

from  openpyxl.chart import BarChart, Reference #to add chart



                                                                  # cell = sheet['a1'] for demo
   
def process_workbook(filename):          #used func to store more data                                                        # cell = sheet.cell(1,1)
    wb = xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']
    
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # print(cell.value)                  #it print the price which is row 3
        
        corrected_price = float(cell.value )* 0.9  # updated the value 
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
        
        
    values = Reference(sheet, min_row = 2, max_row =sheet.max_row,
            min_col=4,
            max_col=4) # to select range of value

    chart =BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)